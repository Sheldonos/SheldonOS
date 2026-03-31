
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# --- File Paths ---
csv_path = '/home/ubuntu/miami_leads/Miami_Cultural_Website_Leads_500.csv'
xlsx_path = '/home/ubuntu/miami_leads/Miami_Cultural_Website_Leads_500.xlsx'

# --- Load Data ---
df = pd.read_csv(csv_path)

# --- Create Workbook and Sheets ---
wb = Workbook()
ws_db = wb.active
ws_db.title = "Master Lead Database"
ws_dash = wb.create_sheet("Dashboard")

# --- 1. Master Lead Database Sheet ---
for r in dataframe_to_rows(df, index=False, header=True):
    ws_db.append(r)

# Auto-fit columns
for col in ws_db.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) if max_length < 50 else 50
    ws_db.column_dimensions[column].width = adjusted_width

# --- 2. Dashboard Sheet ---
def style_cell(cell, text, bold=True, size=14, fill="DDEBF7"):
    cell.value = text
    cell.font = Font(bold=bold, size=size)
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# A. Priority Tier Distribution
style_cell(ws_dash['B2'], 'Priority Tier Distribution', size=16)
tier_counts = df['priority_tier'].value_counts().reset_index()
tier_counts.columns = ['Tier', 'Count']
for row in dataframe_to_rows(tier_counts, index=False, header=True):
    ws_dash.append(row)

# Pie Chart for Tiers
pie = PieChart()
labels = Reference(ws_dash, min_col=1, min_row=2, max_row=len(tier_counts)+1)
data = Reference(ws_dash, min_col=2, min_row=1, max_row=len(tier_counts)+1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Leads by Priority Tier"
ws_dash.add_chart(pie, "D2")

# B. Top Cultural Categories
style_cell(ws_dash['B10'], 'Top 10 Cultural Categories', size=16)
category_counts = df['category_consolidated'].value_counts().nlargest(10).reset_index()
category_counts.columns = ['Category', 'Count']
cat_start_row = ws_dash.max_row + 1
for row in dataframe_to_rows(category_counts, index=False, header=True):
    ws_dash.append(row)

# Bar Chart for Categories
bar = BarChart()
data = Reference(ws_dash, min_col=2, min_row=cat_start_row, max_row=cat_start_row + len(category_counts))
cats = Reference(ws_dash, min_col=1, min_row=cat_start_row + 1, max_row=cat_start_row + len(category_counts))
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.title = "Leads by Cultural Category"
bar.legend = None
ws_dash.add_chart(bar, "D10")

# --- Save Workbook ---
wb.save(xlsx_path)

print(f"Successfully created Cultural leads Excel workbook at: {xlsx_path}")
