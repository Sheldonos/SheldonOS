'''
This script creates a comprehensive Excel workbook for the Miami Website Development leads.
It includes:
1.  A master database of all leads.
2.  A dashboard with visualizations for lead type, category breakdown, and revenue.
3.  A sheet with pre-written outreach email templates.
4.  A pricing guide for different website packages.
'''
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# --- File Paths ---
csv_path = '/home/ubuntu/miami_leads/Miami_Website_Leads_Master.csv'
xlsx_path = '/home/ubuntu/miami_leads/Miami_Website_Leads_Master.xlsx'

# --- Load Data ---
df = pd.read_csv(csv_path)

# --- Create Workbook and Sheets ---
wb = Workbook()
ws_db = wb.active
ws_db.title = "Master Lead Database"
ws_dash = wb.create_sheet("Dashboard")
ws_templates = wb.create_sheet("Outreach Templates")
ws_pricing = wb.create_sheet("Pricing Guide")

# --- 1. Master Lead Database Sheet ---
for r in dataframe_to_rows(df, index=False, header=True):
    ws_db.append(r)

# Auto-fit columns for the database sheet
for col in ws_db.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_db.column_dimensions[column].width = adjusted_width

# --- 2. Dashboard Sheet ---
def style_cell(cell, text, bold=True, size=14, fill="DDEBF7"):
    cell.value = text
    cell.font = Font(bold=bold, size=size)
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# A. Lead Type Distribution
style_cell(ws_dash['B2'], 'Lead Type Distribution', size=16)
lead_type_counts = df['Lead Type'].value_counts().reset_index()
lead_type_counts.columns = ['Lead Type', 'Count']
# Start appending from a safe row
for i, row in enumerate(dataframe_to_rows(lead_type_counts, index=False, header=True)):
    ws_dash.append(row)

# Pie Chart for Lead Types
pie = PieChart()
labels = Reference(ws_dash, min_col=1, min_row=2, max_row=len(lead_type_counts)+1)
data = Reference(ws_dash, min_col=2, min_row=1, max_row=len(lead_type_counts)+1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Lead Type Distribution"
ws_dash.add_chart(pie, "D2")

# B. Category Breakdown
style_cell(ws_dash['B10'], 'Top 5 Categories', size=16)
category_counts = df['Category'].value_counts().nlargest(5).reset_index()
category_counts.columns = ['Category', 'Count']
# Append data for the bar chart
cat_start_row = 11
for row in dataframe_to_rows(category_counts, index=False, header=True):
    ws_dash.append(row)

# Bar Chart for Categories
bar = BarChart()
data = Reference(ws_dash, min_col=2, min_row=cat_start_row, max_row=cat_start_row + len(category_counts))
cats = Reference(ws_dash, min_col=1, min_row=cat_start_row + 1, max_row=cat_start_row + len(category_counts))
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.title = "Leads by Category"
bar.legend = None
ws_dash.add_chart(bar, "D10")

# C. Revenue Opportunity
style_cell(ws_dash['B20'], 'Revenue Opportunity', size=16)

# Clean and convert revenue column to numeric
df['Est Monthly Revenue Loss ($)'] = pd.to_numeric(df['Est Monthly Revenue Loss ($)'].replace('[\\$,]', '', regex=True), errors='coerce').fillna(0)
total_loss = df['Est Monthly Revenue Loss ($)'].sum()
avg_loss = df['Est Monthly Revenue Loss ($)'].mean()

ws_dash['B21'] = 'Total Estimated Monthly Revenue Loss'
ws_dash['C21'] = f"${total_loss:,.2f}"
ws_dash['B22'] = 'Average Loss Per Lead'
ws_dash['C22'] = f"${avg_loss:,.2f}"

# --- 3. Outreach Templates Sheet ---
style_cell(ws_templates['B2'], 'Cold Email Outreach Templates for Miami Market', size=16)

ws_templates['B4'] = "Template 1: For 'No Website' Leads"
ws_templates['B5'] = "Subject: Your [Industry] business in Miami deserves a digital storefront"
ws_templates['B6'] = '''Hi [Decision Maker Name],\n\nMy name is [Your Name], and I specialize in creating professional websites for businesses in Miami.\n\nI noticed that [Business Name] doesn\'t currently have a website. In today\'s digital-first world, that means missing out on countless customers who are searching online for [services you offer] right now in [Neighborhood].\n\nA professional website could help you:\n- Attract new customers searching on Google\n- Showcase your work and build credibility\n- Allow customers to easily contact you or book appointments online\n\nI have some ideas specific to the [Industry] space that I\'d love to share. Would you be open to a brief 15-minute chat next week?\n\nBest,\n[Your Name]'''
ws_templates.row_dimensions[6].height = 200

ws_templates['B8'] = "Template 2: For 'Outdated Website' Leads"
ws_templates['B9'] = "Subject: A quick thought on [Business Name]'s website"
ws_templates['B10'] = '''Hi [Decision Maker Name],\n\nI was looking for [Industry] services in Miami and came across your website. It\'s great that you have an online presence!\n\nI noticed the design is a bit dated and doesn\'t work well on mobile devices. A modern, mobile-friendly website can significantly boost customer engagement and conversions.\n\nWe help businesses like yours refresh their online presence to better reflect the quality of their work and attract more clients. I have a few specific suggestions for your site that could make a big impact.\n\nWould you have 15 minutes to discuss them sometime next week?\n\nBest,\n[Your Name]'''
ws_templates.row_dimensions[10].height = 200

# --- 4. Pricing Guide Sheet ---
style_cell(ws_pricing['B2'], 'Website Development Pricing Guide - Miami Market', size=16)

headers = ['Package', 'Price Range (USD)', 'Best For', 'Key Features']
pricing_data = [
    ('Starter', '$1,500 - $3,000', 'New businesses or those with no website.', '3-5 page informational site, Contact form, Basic SEO, Mobile responsive'),
    ('Professional', '$3,000 - $7,500', 'Established businesses with an outdated site.', '5-10 pages, CMS (e.g., WordPress), Blog, Advanced SEO, Online booking/e-commerce integration'),
    ('Premium', '$7,500 - $15,000+', 'Businesses needing a full redesign with custom features.', 'Custom design & development, Advanced functionality, API integrations, Full e-commerce, Content strategy'),
]

ws_pricing.append(headers)
for row in pricing_data:
    ws_pricing.append(row)

# Style the pricing table
for row in ws_pricing['A3:D6']:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# --- Save Workbook ---
wb.save(xlsx_path)

print(f"Successfully created Excel workbook at: {xlsx_path}")
