'''
This script creates a comprehensive Excel workbook for the Miami AI Services leads.
It includes:
1.  A master database of all AI leads.
2.  A dashboard with visualizations for revenue tiers, industry breakdown, and pipeline value.
3.  A pricing and solution guide for the AI stack.
'''
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# --- File Paths ---
csv_path = '/home/ubuntu/miami_leads/Miami_AI_Services_Leads_FINAL.csv'
xlsx_path = '/home/ubuntu/miami_leads/Miami_AI_Services_Leads_MASTER.xlsx'

# --- Load Data ---
df = pd.read_csv(csv_path)

# --- Data Cleaning & Preparation for Dashboard ---
def get_revenue_tier(potential):
    try:
        # Extract the first number from the potential string
        nums = re.findall(r'[\d,]+', str(potential).replace('$', ''))
        if not nums:
            return 'Starter'
        min_rev = int(nums[0].replace(',', ''))
        if min_rev < 5000:
            return 'Starter'
        elif 5000 <= min_rev < 15000:
            return 'Growth'
        elif 15000 <= min_rev < 30000:
            return 'Professional'
        else:
            return 'Enterprise'
    except:
        return 'Starter'

df['Revenue Tier'] = df['Revenue Potential'].apply(get_revenue_tier)

# --- Create Workbook and Sheets ---
wb = Workbook()
ws_db = wb.active
ws_db.title = "Master Lead Database"
ws_dash = wb.create_sheet("Dashboard")
ws_guide = wb.create_sheet("Solution & Pricing Guide")

# --- 1. Master Lead Database Sheet ---
for r in dataframe_to_rows(df, index=False, header=True):
    ws_db.append(r)

# Auto-fit columns
for col in ws_db.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) if max_length < 50 else 50
    ws_db.column_dimensions[column].width = adjusted_width

# --- 2. Dashboard Sheet ---
def style_cell(cell, text, bold=True, size=14, fill="DDEBF7"):
    cell.value = text
    cell.font = Font(bold=bold, size=size)
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# A. Revenue Tier Breakdown
style_cell(ws_dash['B2'], 'Revenue Tier Breakdown', size=16)
tier_counts = df['Revenue Tier'].value_counts().reset_index()
tier_counts.columns = ['Tier', 'Count']
for row in dataframe_to_rows(tier_counts, index=False, header=True):
    ws_dash.append(row)

# Pie Chart for Tiers
pie = PieChart()
labels = Reference(ws_dash, min_col=1, min_row=2, max_row=len(tier_counts)+1)
data = Reference(ws_dash, min_col=2, min_row=1, max_row=len(tier_counts)+1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Leads by Revenue Tier"
ws_dash.add_chart(pie, "D2")

# B. Top Industry Categories
style_cell(ws_dash['B10'], 'Top 10 Industry Categories', size=16)
category_counts = df['Category'].value_counts().nlargest(10).reset_index()
category_counts.columns = ['Category', 'Count']
cat_start_row = ws_dash.max_row + 1
for row in dataframe_to_rows(category_counts, index=False, header=True):
    ws_dash.append(row)

# Bar Chart for Categories
bar = BarChart()
data = Reference(ws_dash, min_col=2, min_row=cat_start_row, max_row=cat_start_row + len(category_counts))
cats = Reference(ws_dash, min_col=1, min_row=cat_start_row + 1, max_row=cat_start_row + len(category_counts))
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.title = "Leads by Category"
bar.legend = None
ws_dash.add_chart(bar, "D10")

# --- 3. Solution & Pricing Guide Sheet ---
style_cell(ws_guide['B2'], 'AI Solution Stack & Pricing - Miami Market', size=16)

# Add content for the guide
guide_content = [
    ["GoHighLevel (GHL) — CRM & Marketing Automation", ""],
    ["GHL Starter", "$1,200/mo - CRM setup, 3 automation workflows, reputation mgmt"],
    ["GHL Growth", "$2,500/mo - Full CRM, 10 workflows, SMS campaigns, lead nurturing"],
    ["GHL Pro", "$4,500/mo - Full stack, AI chatbot, custom integrations"],
    ["", ""],
    ["OpenAI Integration — AI Content & Documentation", ""],
    ["AI Content Starter", "$1,500/mo - Blog/social content automation, basic chatbot"],
    ["AI Documentation", "$3,500/mo - Custom document drafting, contract templates"],
    ["AI Intelligence Suite", "$8,000/mo - Full custom AI system, trained on client data"],
    ["", ""],
    ["Lovabl — Client Engagement & Personalization", ""],
    ["Lovabl Engage", "$1,200/mo - Automated re-engagement campaigns, review requests"],
    ["Lovabl Loyalty", "$2,500/mo - Full loyalty program, VIP tiers, journey mapping"],
    ["", ""],
    ["Kling — Workflow & Operations Automation", ""],
    ["Kling Ops Starter", "$1,500/mo - Scheduling automation, basic dispatching"],
    ["Kling Ops Pro", "$3,500/mo - Full dispatching, job tracking, automated invoicing"],
]

for row_data in guide_content:
    ws_guide.append(row_data)

# --- Save Workbook ---
wb.save(xlsx_path)

print(f"Successfully created AI leads Excel workbook at: {xlsx_path}")
